from decimal import Decimal, InvalidOperation
from datetime import datetime, date, time, timedelta
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, close_old_connections, connection
from django.db.utils import OperationalError, InterfaceError
from django.contrib.auth.models import User
from django.utils import timezone

from patients.models import Patient
from appointments.models import Appointment, Doctor, Service, Treatment
from billing.models import Invoice, InvoiceItem, Payment

try:
    import openpyxl
except ImportError:
    openpyxl = None


SERVICE_NAMES = [
    'BRACES', 'BRACES METALLIC', 'BRACES REVIEW', 'SCALING', 'POLISHING',
    'RCT', 'EXTRACTION', 'WHITENING', 'DENTAL CHECK UP', 'X RAY', 'KIT',
    'REVIEW', 'SP 10', 'SP 6', 'DENTURES'
]

SERVICE_MAP = [
    ('BRACES AND KIT', 'BRACES METALLIC'),
    ('BRACES METALLIC', 'BRACES METALLIC'),
    ('BRACES REVIEW', 'BRACES REVIEW'),
    ('BRACES', 'BRACES'),
    ('SCALING AND POLISHING', 'SCALING'),
    ('SCALING', 'SCALING'),
    ('POLISHING', 'POLISHING'),
    ('RCT', 'RCT'),
    ('EXTRACTION', 'EXTRACTION'),
    ('X RAY', 'X RAY'),
    ('FILLINGS', 'DENTAL CHECK UP'),
    ('FILLING', 'DENTAL CHECK UP'),
    ('CHECK UP', 'DENTAL CHECK UP'),
    ('DENTAL CHECK UP', 'DENTAL CHECK UP'),
    ('SP 10', 'WHITENING'),
    ('SP 6', 'SCALING'),
    ('DENTURES', 'DENTURES'),
    ('RETAINERS', 'BRACES REVIEW'),
    ('REVIEW', 'BRACES REVIEW'),
    ('CROWN', 'DENTAL CHECK UP'),
    ('BRIDGE', 'DENTAL CHECK UP'),
]


def clean_text(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def normalize_name(value):
    return clean_text(value).upper()


def normalize_gender(value):
    value = clean_text(value).lower()
    if value.startswith('m'):
        return 'M'
    if value.startswith('f'):
        return 'F'
    return 'O'


def clean_phone(value):
    if value is None:
        return ''
    raw = clean_text(value)
    if not raw:
        return ''
    parts = re.split(r'[/,;]|\s+and\s+', raw, flags=re.I)
    first = re.sub(r'\D', '', parts[0])
    if first.startswith('256'):
        return '+' + first
    if first.startswith('0') and len(first) >= 9:
        return '+256' + first[1:]
    if first.startswith('7') and len(first) >= 9:
        return '+256' + first
    return first


def split_name(full_name):
    parts = normalize_name(full_name).split()
    if not parts:
        return 'UNKNOWN', ''
    if len(parts) == 1:
        return parts[0], ''
    return parts[0], ' '.join(parts[1:])


def parse_money(value):
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean_text(value).upper().replace(',', '').replace('UGX', '')
    if text in ('', 'NIL', 'NONE', '-', 'N/A'):
        return Decimal('0')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def parse_date(value):
    # The clinic register is DD/MM/YYYY. Excel native date cells in this
    # workbook were saved with month/day components, so 2026-09-06 means
    # 09/06/2026 (9 June 2026) in the register.
    if isinstance(value, datetime):
        try:
            return date(value.year, value.day, value.month)
        except ValueError:
            return None
    if isinstance(value, date):
        try:
            return date(value.year, value.day, value.month)
        except ValueError:
            return None
    text = clean_text(value)
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    # One source row contains 5/06/206; treat a 3-digit year as 2000 + year.
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{3})', text)
    if m:
        d, mo, y = map(int, m.groups())
        if y == 206:
            y = 2026
        elif y < 1000:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def service_for(treatment):
    text = normalize_name(treatment)
    for key, target in SERVICE_MAP:
        if key in text:
            return target
    return 'DENTAL CHECK UP'


def doctor_name_for(raw):
    text = normalize_name(raw)
    if not text:
        return 'UNASSIGNED'
    # If multiple clinicians are written in one cell, retain the first clinician
    # as the responsible clinician and preserve the original in visit notes.
    if ' AND ' in text:
        text = text.split(' AND ', 1)[0].strip()
    text = re.sub(r'^DR\.?\s+', '', text)
    return text


class Command(BaseCommand):
    help = 'Import the Dora\'s Dental Gem REGISTER.xlsx historical register.'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', default='REGISTER.xlsx')
        parser.add_argument('--dry-run', action='store_true', help='Validate and report without writing data.')
        parser.add_argument('--no-visits', action='store_true', help='Import patients and billing, but do not create historical visits.')
        parser.add_argument('--no-billing', action='store_true', help='Import patients and visits, but do not create invoices/payments.')
        parser.add_argument('--created-by', default='', help='Username to record as creator of imported records.')
        parser.add_argument('--limit', type=int, default=0, help='Import only N parsed patient/visit rows (0 = all).')
        parser.add_argument('--start', type=int, default=1, help='Start at this parsed patient/visit row (1-based).')
        parser.add_argument('--batch-size', type=int, default=500, help='Rows per database transaction (default: 500).')

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl is required. Install it with: python -m pip install openpyxl')

        source = Path(options['file']).expanduser()
        if not source.exists():
            raise CommandError(f'Excel file not found: {source}')

        wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
        if 'daily record' not in wb.sheetnames:
            raise CommandError("Workbook must contain a 'daily record' sheet.")
        ws = wb['daily record']

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('The daily record sheet is empty.')

        creator = None
        if options['created_by']:
            creator = User.objects.filter(username=options['created_by']).first()
            if not creator:
                raise CommandError(f"User '{options['created_by']}' does not exist.")

        stats = {k: 0 for k in (
            'rows', 'patients_created', 'patients_reused', 'visits_created',
            'treatments_created', 'invoices_created', 'payments_created',
            'doctors_created', 'services_created', 'warnings'
        )}
        warnings = []
        parsed = []
        current_date = None

        for excel_row, row in enumerate(rows[1:], start=2):
            values = list(row) + [None] * max(0, 10 - len(row))
            date_cell, name, gender, treatment, address, offer, paid, balance, doctor, contact = values[:10]

            # Date-only row: carry the date forward to following patient rows.
            if date_cell is not None and all(v in (None, '') for v in values[1:]):
                parsed_date = parse_date(date_cell)
                if parsed_date:
                    current_date = parsed_date
                else:
                    warnings.append(f'Row {excel_row}: could not parse date {date_cell!r}')
                continue

            name = clean_text(name)
            if not name:
                continue

            stats['rows'] += 1
            visit_date = current_date or parse_date(date_cell)
            if current_date is None and not parse_date(date_cell):
                warnings.append(f'Row {excel_row}: patient {name!r} has no usable date; row skipped.')

            if visit_date is None:
                continue

            parsed.append({
                'row': excel_row,
                'date': visit_date,
                'name': name,
                'gender': normalize_gender(gender),
                'treatment': clean_text(treatment),
                'address': clean_text(address),
                'offer': clean_text(offer),
                'paid': parse_money(paid),
                'balance': parse_money(balance),
                'doctor_raw': clean_text(doctor),
                'doctor': doctor_name_for(doctor),
                'phone': clean_phone(contact),
                'contact_raw': clean_text(contact),
            })

        # Build the patient's true registration date from the ENTIRE sheet
        # before applying --limit. Registration date is the earliest date on
        # which that patient appears anywhere in the register.
        first_date_by_name = {}
        for item in parsed:
            patient_key = normalize_name(item['name'])
            if item['date'] and (
                patient_key not in first_date_by_name
                or item['date'] < first_date_by_name[patient_key]
            ):
                first_date_by_name[patient_key] = item['date']

        start = max(1, int(options.get('start', 1)))
        if start > 1:
            parsed = parsed[start - 1:]
        if options['limit'] and options['limit'] > 0:
            parsed = parsed[:options['limit']]

        if options['dry_run']:
            self.stdout.write(self.style.SUCCESS(f'Validated {len(parsed)} patient/visit rows.'))
            self.stdout.write(f'Date groups: {len({x["date"] for x in parsed})}')
            self.stdout.write(f'Rows with phone: {sum(bool(x["phone"]) for x in parsed)}')
            self.stdout.write(f'Rows with balance: {sum(x["balance"] > 0 for x in parsed)}')
            if warnings:
                self.stdout.write(self.style.WARNING(f'Warnings: {len(warnings)}'))
                for w in warnings[:20]:
                    self.stdout.write(f'  - {w}')
            return

        batch_size = max(1, options['batch_size'])
        total = len(parsed)
        if total == 0:
            self.stdout.write(self.style.WARNING('No patient/visit rows to import.'))
            return

        # Cache static lookup data across batches, but keep transactions short.
        service_cache = {}
        doctor_cache = {}

        def prepare_lookups():
            close_old_connections()
            for service_name in SERVICE_NAMES:
                service = Service.objects.filter(name=service_name).first()
                if service is None:
                    service = Service.objects.create(
                        name=service_name,
                        description=service_name,
                        price=Decimal('0'),
                        duration_minutes=30,
                        is_active=True,
                    )
                    stats['services_created'] += 1
                service_cache[service_name] = service

        prepare_lookups()

        # Preload lookup data once. The register is only 1,854 rows, so this
        # dramatically reduces PostgreSQL round-trips while keeping transactions short.
        patient_by_phone = {}
        patient_by_name = {}
        for p in Patient.objects.all().only('id','first_name','last_name','phone','address','gender','registered_at','reason_for_visit'):
            key = normalize_name(f"{p.first_name} {p.last_name}")
            patient_by_name.setdefault(key, p)
            if p.phone:
                patient_by_phone.setdefault(p.phone, p)
                if p.phone.startswith('+256'):
                    patient_by_phone.setdefault('0' + p.phone[4:], p)

        for service in Service.objects.all():
            service_cache[service.name] = service
        for doctor in Doctor.objects.all():
            doctor_cache[doctor.name] = doctor

        imported_appointments = {}
        for a in Appointment.objects.filter(notes__startswith='Imported from REGISTER.xlsx row ').select_related('patient','doctor','service'):
            m = re.search(r'Imported from REGISTER\.xlsx row (\d+)\.', a.notes or '')
            if m:
                imported_appointments[int(m.group(1))] = a

        imported_invoices = {}
        for inv in Invoice.objects.filter(notes__startswith='Imported from REGISTER.xlsx row '):
            m = re.search(r'Imported from REGISTER\.xlsx row (\d+)\.', inv.notes or '')
            if m:
                imported_invoices[int(m.group(1))] = inv

        appointment_slot_counts = {}

        def import_batch(batch):
            nonlocal patient_by_phone, patient_by_name, service_cache, doctor_cache
            close_old_connections()
            with transaction.atomic():
                for item in batch:
                    first, last = split_name(item['name'])
                    phone = item['phone']
                    name_key = normalize_name(item['name'])

                    patient = patient_by_phone.get(phone) if phone else None
                    if patient is None:
                        patient = patient_by_name.get(name_key)

                    if patient is None:
                        patient = Patient.objects.create(
                            first_name=first, last_name=last, gender=item['gender'],
                            phone=phone, address=item['address'],
                            reason_for_visit=item['treatment'],
                        )
                        stats['patients_created'] += 1
                    else:
                        stats['patients_reused'] += 1
                        changed = []
                        if item['address'] and not patient.address:
                            patient.address = item['address']; changed.append('address')
                        if phone and not patient.phone:
                            patient.phone = phone; changed.append('phone')
                        if item['gender'] and patient.gender == 'O' and item['gender'] != 'O':
                            patient.gender = item['gender']; changed.append('gender')
                        if changed:
                            patient.save(update_fields=changed)

                    # Registration date is the earliest appearance in the entire register.
                    historical_date = first_date_by_name.get(name_key)
                    if historical_date:
                        historical_first = timezone.make_aware(
                            datetime.combine(historical_date, time.min),
                            timezone.get_current_timezone(),
                        )
                        existing = patient.registered_at
                        if not existing or historical_first < existing:
                            patient.registered_at = historical_first
                            patient.save(update_fields=['registered_at'])

                    patient_by_name[name_key] = patient
                    if phone:
                        patient_by_phone.setdefault(phone, patient)

                    service_name = service_for(item['treatment'])
                    service = service_cache.get(service_name)
                    if service is None:
                        service = Service.objects.filter(name=service_name).first()
                        if service is None:
                            service = Service.objects.create(
                                name=service_name, description=service_name,
                                price=Decimal('0'), duration_minutes=30, is_active=True,
                            )
                            stats['services_created'] += 1
                        service_cache[service_name] = service

                    doctor = doctor_cache.get(item['doctor'])
                    if doctor is None:
                        doctor = Doctor.objects.filter(name=item['doctor']).first()
                        if doctor is None:
                            doctor = Doctor.objects.create(name=item['doctor'], is_active=True)
                            stats['doctors_created'] += 1
                        doctor_cache[item['doctor']] = doctor

                    appointment = imported_appointments.get(item['row'])
                    marker = f"Imported from REGISTER.xlsx row {item['row']}."

                    if not options['no_visits']:
                        if appointment is not None:
                            changed = []
                            if appointment.appointment_date != item['date']:
                                appointment.appointment_date = item['date']; changed.append('appointment_date')
                            if appointment.patient_id != patient.id:
                                appointment.patient = patient; changed.append('patient')
                            if changed:
                                appointment.save(update_fields=changed)
                        else:
                            slot_key = (doctor.id, item['date'])
                            if slot_key not in appointment_slot_counts:
                                appointment_slot_counts[slot_key] = Appointment.objects.filter(
                                    doctor=doctor, appointment_date=item['date']
                                ).count()
                            n = appointment_slot_counts[slot_key]
                            appointment_slot_counts[slot_key] += 1
                            appointment_time = (datetime.combine(item['date'], time.min) + timedelta(minutes=n)).time()
                            notes = marker + f" Original doctor: {item['doctor_raw'] or 'N/A'}."
                            if item.get('contact_raw') and item['contact_raw'] != item['phone']:
                                notes += f" Original register contact: {item['contact_raw']}."
                            if item['offer']:
                                notes += f" Register status/offer: {item['offer']}."
                            appointment = Appointment.objects.create(
                                patient=patient, doctor=doctor, service=service,
                                appointment_date=item['date'], appointment_time=appointment_time,
                                duration_minutes=service.duration_minutes or 30, status='completed',
                                notes=notes, consultation_notes=item['treatment'],
                                treatment_cost=item['paid'] + item['balance'],
                                notification_phone=patient.phone or '', created_by=creator,
                            )
                            imported_appointments[item['row']] = appointment
                            stats['visits_created'] += 1

                        Treatment.objects.get_or_create(
                            patient=patient, doctor=doctor, service=service, appointment=appointment,
                            defaults={'notes': marker + ' ' + (item['treatment'] or service.name),
                                      'amount': item['paid'] + item['balance']},
                        )

                    if not options['no_billing']:
                        total_amount = item['paid'] + item['balance']
                        if total_amount > 0:
                            invoice_number = f"REG-{item['row']:05d}"
                            invoice = imported_invoices.get(item['row'])
                            if invoice is None:
                                # Verify the appointment is actually persisted before using it as FK.
                                persisted_appointment = None
                                if appointment is not None and appointment.pk:
                                    persisted_appointment = Appointment.objects.filter(pk=appointment.pk).first()
                                invoice = Invoice.objects.create(
                                    invoice_number=invoice_number, patient=patient,
                                    appointment=persisted_appointment, patient_name=patient.full_name,
                                    patient_phone=patient.phone or '', issue_date=item['date'],
                                    due_date=item['date'], subtotal=total_amount, tax_rate=Decimal('0'),
                                    tax_amount=Decimal('0'), discount=Decimal('0'), total_amount=total_amount,
                                    amount_paid=item['paid'], balance_due=item['balance'],
                                    status='paid' if item['balance'] <= 0 else ('partially_paid' if item['paid'] > 0 else 'sent'),
                                    payment_method='cash' if item['paid'] > 0 else '',
                                    notes=marker + f" Treatment: {item['treatment']}.",
                                    created_by=creator.username if creator else '',
                                    updated_by=creator.username if creator else '',
                                )
                                imported_invoices[item['row']] = invoice
                                stats['invoices_created'] += 1
                                InvoiceItem.objects.create(
                                    invoice=invoice, service=service, description=item['treatment'] or service.name,
                                    quantity=1, unit_price=total_amount, total_price=total_amount,
                                )
                                if item['paid'] > 0:
                                    Payment.objects.create(
                                        invoice=invoice, amount=item['paid'], payment_date=item['date'],
                                        payment_method='cash', status='completed', notes=marker,
                                        processed_by=creator.username if creator else '',
                                    )
                                    stats['payments_created'] += 1

                processed = 0
        processed = 0


        for start_idx in range(0, total, batch_size):
            batch = parsed[start_idx:start_idx + batch_size]
            last_error = None
            for attempt in range(1, 4):
                try:
                    import_batch(batch)
                    last_error = None
                    break
                except (OperationalError, InterfaceError) as exc:
                    last_error = exc
                    warnings.append(
                        f"Batch {start_idx + 1}-{start_idx + len(batch)} database connection error; retry {attempt}/3."
                    )
                    try:
                        connection.close()
                    except Exception:
                        pass
                    close_old_connections()
                    if attempt < 3:
                        import time as _time
                        _time.sleep(2 * attempt)
            if last_error is not None:
                raise CommandError(
                    f"Database connection failed after 3 attempts for rows "
                    f"{start_idx + 1}-{start_idx + len(batch)}: {last_error}"
                ) from last_error
            processed += len(batch)
            self.stdout.write(f"Imported {processed}/{total} rows...")

        self.stdout.write(self.style.SUCCESS('\nREGISTER IMPORT COMPLETE'))
        for key, value in stats.items():
            self.stdout.write(f'{key.replace("_", " ").title()}: {value}')
        if warnings:
            self.stdout.write(self.style.WARNING(f'Warnings: {len(warnings)}'))
            for w in warnings[:20]:
                self.stdout.write(f'  - {w}')
